#!/usr/bin/env python3
"""
Convert CIA data.xlsx into synthesis.md (markdown digest of non-empty sheets).

Usage:  scripts/cia-synthesize.py projects/<slug>/cia/

Reads:  data.xlsx
Writes: synthesis.md  (overwrites)

The output is what server/contentos.js readCiaData() injects into Step 1+2
LLM prompts as primary-source market data.
"""
import sys
import os
from pathlib import Path
from openpyxl import load_workbook

MAX_ROWS_PER_SHEET = 50
MAX_COLS = 12
MAX_CELL_LEN = 200


def _strip_non_bmp(s):
    # Strip non-BMP code points (4-byte UTF-8 / emojis) — some upstream JSON
    # parsers (seen in Anthropic via flatkey CF proxy) reject long bodies
    # containing them with "invalid escaped character".
    return "".join(c for c in s if ord(c) < 0x10000)


def cell_str(v):
    if v is None:
        return ""
    s = str(v).replace("|", "/").replace("\n", " ").strip()
    s = _strip_non_bmp(s)
    return s[:MAX_CELL_LEN] + "…" if len(s) > MAX_CELL_LEN else s


def sheet_to_md(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows or all(all(c is None or str(c).strip() == "" for c in r) for r in rows):
        return None
    header = [cell_str(c) for c in rows[0][:MAX_COLS]]
    body = rows[1:MAX_ROWS_PER_SHEET + 1]
    if not header or all(h == "" for h in header):
        return None
    lines = ["| " + " | ".join(header) + " |", "| " + " | ".join(["---"] * len(header)) + " |"]
    for r in body:
        cells = [cell_str(c) for c in r[:MAX_COLS]]
        cells += [""] * (len(header) - len(cells))
        lines.append("| " + " | ".join(cells) + " |")
    n_total = len(rows) - 1
    if n_total > MAX_ROWS_PER_SHEET:
        lines.append(f"\n_(showing {MAX_ROWS_PER_SHEET}/{n_total} rows — full data in data.xlsx)_")
    return "\n".join(lines)


def main():
    if len(sys.argv) != 2:
        print("Usage: cia-synthesize.py <cia-dir>", file=sys.stderr)
        sys.exit(1)
    cia_dir = Path(sys.argv[1]).resolve()
    xlsx = cia_dir / "data.xlsx"
    out = cia_dir / "synthesis.md"
    if not xlsx.exists():
        print(f"data.xlsx not found in {cia_dir}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    parts = [f"# CIA Synthesis — {cia_dir.parent.name}", "",
             "Auto-generated from `data.xlsx` by `scripts/cia-synthesize.py`. "
             "Source: CIA-insight pipeline (Ahrefs / DataForSEO / Apify TikTok+Reddit / iTunes / YouTube).",
             ""]
    used = []
    for name in wb.sheetnames:
        ws = wb[name]
        md = sheet_to_md(ws)
        if md is None:
            continue
        parts.append(f"## {name}")
        parts.append("")
        parts.append(md)
        parts.append("")
        used.append(name)
    if not used:
        parts.append("_(No non-empty sheets — CIA fetches returned no rows.)_")
    else:
        parts.insert(4, f"**Non-empty sheets**: {', '.join(used)}")
        parts.insert(5, "")
    out.write_text("\n".join(parts))
    sz = out.stat().st_size
    print(f"✓ wrote {out} ({sz} bytes, {len(used)} sheets)")


if __name__ == "__main__":
    main()
